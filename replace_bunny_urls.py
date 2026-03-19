#!/usr/bin/env python3
"""
Replace Bunny CDN image URLs in the product database Excel file
with new S3 image URLs from image_urls.txt.

Matching strategy:
  1. Exact match: Part_Number == SKU folder name
  2. Last-token match: last word of Part_Number == SKU folder name
  3. No match: leave Bunny_Image_URLs unchanged, log warning

Output: new Excel file with updated Bunny_Image_URLs column.
"""

import os
import re
from datetime import date
from pathlib import PurePosixPath
from collections import defaultdict

import openpyxl

# ── Configuration ──────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
IMAGE_URLS_FILE = os.path.join(BASE_DIR, "image_urls.txt")
EXCEL_INPUT = os.path.join(
    BASE_DIR,
    "complete_product_urls_database_with_descriptions_enriched_bunny_fixed_updated_20-03-2026.xlsx",
)
EXCEL_OUTPUT = os.path.join(
    BASE_DIR,
    f"updated_product_urls_database_{date.today().isoformat()}.xlsx",
)

BUNNY_COL_INDEX = 7  # 1-based column index for Bunny_Image_URLs (col G = 7)
PART_COL_INDEX = 1   # 1-based column index for Part_Number (col A = 1)


def natural_sort_key(url: str) -> tuple:
    """Sort URLs by their filename so 01_... comes before 02_... etc."""
    filename = url.rsplit("/", 1)[-1]
    # Extract leading number for numeric sort
    match = re.match(r"(\d+)", filename)
    num = int(match.group(1)) if match else 999
    return (num, filename)


def build_sku_url_map(image_urls_path: str) -> dict[str, list[str]]:
    """
    Parse image_urls.txt and build a dict: SKU -> [sorted list of URLs].
    SKU is extracted as the parent folder of each image filename.
    """
    sku_map: dict[str, list[str]] = defaultdict(list)

    with open(image_urls_path, "r") as f:
        for line in f:
            url = line.strip()
            if not url:
                continue
            path = PurePosixPath(url)
            sku = path.parent.name
            if sku:
                sku_map[sku].append(url)

    # Sort each SKU's URLs by filename (01_, 02_, etc.)
    for sku in sku_map:
        sku_map[sku].sort(key=natural_sort_key)

    return dict(sku_map)


def find_urls_for_part(part_number: str, sku_map: dict[str, list[str]]) -> tuple[list[str], str]:
    """
    Try to find matching URLs for a given Part_Number.
    Returns (urls, match_type) where match_type is 'exact', 'last-token', or 'none'.
    """
    part = part_number.strip()

    # 1. Exact match
    if part in sku_map:
        return sku_map[part], "exact"

    # 2. Last-token match (e.g., "TL-ARCHER AX10" -> "AX10")
    tokens = part.split()
    if len(tokens) > 1:
        last_token = tokens[-1]
        if last_token in sku_map:
            return sku_map[last_token], "last-token"

    # 3. No match
    return [], "none"


def main():
    print("=" * 70)
    print("  Bunny URL Replacement Tool")
    print("=" * 70)

    # ── Step 1: Build SKU → URL map ───────────────────────────────────────
    print(f"\n[1/4] Reading image URLs from: {os.path.basename(IMAGE_URLS_FILE)}")
    sku_map = build_sku_url_map(IMAGE_URLS_FILE)
    total_urls = sum(len(urls) for urls in sku_map.values())
    print(f"       Found {total_urls:,} URLs across {len(sku_map)} unique SKU folders")

    # ── Step 2: Load Excel ────────────────────────────────────────────────
    print(f"\n[2/4] Loading Excel: {os.path.basename(EXCEL_INPUT)}")
    wb = openpyxl.load_workbook(EXCEL_INPUT)
    ws = wb.active
    total_rows = ws.max_row - 1  # Exclude header
    print(f"       Found {total_rows} data rows, {ws.max_column} columns")

    # ── Step 3: Process each row ──────────────────────────────────────────
    print(f"\n[3/4] Matching and replacing URLs...")

    stats = {
        "exact": 0,
        "last-token": 0,
        "none": 0,
        "urls_replaced": 0,
        "empty_part": 0,
    }
    unmatched = []
    ambiguous = []
    last_token_matches = []

    # Track which last-token SKUs map to multiple Part_Numbers
    last_token_usage: dict[str, list[str]] = defaultdict(list)

    for row_idx in range(2, ws.max_row + 1):
        part_cell = ws.cell(row=row_idx, column=PART_COL_INDEX)
        bunny_cell = ws.cell(row=row_idx, column=BUNNY_COL_INDEX)

        part_number = str(part_cell.value).strip() if part_cell.value else ""

        if not part_number:
            stats["empty_part"] += 1
            continue

        urls, match_type = find_urls_for_part(part_number, sku_map)

        if match_type == "none":
            stats["none"] += 1
            unmatched.append(part_number)
        else:
            stats[match_type] += 1
            new_value = "|".join(urls)
            bunny_cell.value = new_value
            stats["urls_replaced"] += len(urls)

            if match_type == "last-token":
                token = part_number.split()[-1]
                last_token_usage[token].append(part_number)
                last_token_matches.append((part_number, token, len(urls)))

    # Detect ambiguous mappings (multiple Part_Numbers → same SKU folder)
    for token, parts in last_token_usage.items():
        if len(parts) > 1:
            ambiguous.append((token, parts))

    # ── Step 4: Save output ───────────────────────────────────────────────
    print(f"\n[4/4] Saving to: {os.path.basename(EXCEL_OUTPUT)}")
    wb.save(EXCEL_OUTPUT)
    wb.close()
    print(f"       Saved successfully!")

    # ── Summary Report ────────────────────────────────────────────────────
    print("\n" + "=" * 70)
    print("  SUMMARY REPORT")
    print("=" * 70)
    print(f"\n  Total data rows:         {total_rows}")
    print(f"  Exact matches:           {stats['exact']}")
    print(f"  Last-token matches:      {stats['last-token']}")
    print(f"  Total matched:           {stats['exact'] + stats['last-token']}")
    print(f"  No match (unchanged):    {stats['none']}")
    print(f"  Empty Part_Number rows:  {stats['empty_part']}")
    print(f"  Total URLs written:      {stats['urls_replaced']:,}")

    if last_token_matches:
        print(f"\n  ── Last-Token Matches ({len(last_token_matches)}) ──")
        for part, token, count in sorted(last_token_matches):
            print(f"    {part:<30s} → folder '{token}' ({count} URLs)")

    if ambiguous:
        print(f"\n  ⚠ AMBIGUOUS MAPPINGS (multiple parts → same folder) ──")
        print(f"    These share the same image set — review manually:")
        for token, parts in sorted(ambiguous):
            print(f"    Folder '{token}': {', '.join(parts)}")

    if unmatched:
        print(f"\n  ✗ UNMATCHED PARTS ({len(unmatched)}) — Bunny URLs left unchanged ──")
        for part in sorted(unmatched):
            print(f"    - {part}")

    print("\n" + "=" * 70)
    print(f"  Output file: {EXCEL_OUTPUT}")
    print("=" * 70)


if __name__ == "__main__":
    main()
